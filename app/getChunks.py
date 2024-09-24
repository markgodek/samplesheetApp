import openpyxl, os

def get_chunks(sheet):
    max_row = sheet.max_row
    row_index = 0
    ID = None

    chunks_dict = {}
    chunks_list = []
    empty_row_counter = 0
    new_chunk = openpyxl.Workbook().create_sheet()

    #read 14 rows, discard 2 rows
    for row in sheet.iter_rows(values_only=True):
        row_index += 1
        # when we find the plate ID, store it in the chunks_dict
        if row[0] == 'Plate ID:':
            ID = row[1]
        # after seeing the 2 rows we don't want between the metadata and table, add rows to the chunk
        if empty_row_counter == 2 and row[0] is not None:
            new_chunk.append(row[1:len(row)])
        # if we've seen 4 empty rows, the chunk is finished
        # append the chunk to the list and initialize the next chunk
        if empty_row_counter == 4:
            chunks_dict[ID] = new_chunk
            chunks_list.append(new_chunk)  # append the constructed chunk to chunks to return
            new_chunk = openpyxl.Workbook().create_sheet()  # reset chunk to return
            empty_row_counter = 0
        if row[0] is None:
            empty_row_counter += 1
        # Store the last chunk after iterating over all the rows
        if row_index == max_row:
            chunks_dict[ID] = new_chunk
            chunks_list.append(new_chunk)  # Append the last chunk if it exists
    print(chunks_dict.keys())
    return chunks_dict


def get_empty_wells(path):
    plate = {}
    rowNames = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    # Load the workbook
    workbook = openpyxl.load_workbook(path)

    # Select the active worksheet (or specify a worksheet by name)
    sheet = workbook.active

    chunks = get_chunks(sheet)

base_path = base_path = os.path.abspath(".")
path = os.path.join(base_path, 'full_platemap.xlsx')
get_empty_wells(path)