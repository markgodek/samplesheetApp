import openpyxl


def get_plates(sheet):
    max_row = sheet.max_row
    row_index = 0
    ID = None

    plates = {}
    empty_row_counter = 0
    new_plate = openpyxl.Workbook().create_sheet()

    # Delete columns N (14) and beyond
    if sheet.max_column >= 14:  # Check if there are any columns to delete
        sheet.delete_cols(14, sheet.max_column - 13)  # Delete from column 14 (N) onward

    #read 14 rows, discard 2 rows
    for row in sheet.iter_rows(values_only=True):
        row_index += 1
        # store the plate ID when we find it
        if row[0] == 'Plate ID:':
            ID = row[1]
        # after seeing the 2 rows we don't want between the metadata and table, add rows to the plate
        if empty_row_counter == 2 and row[0] is not None:
            new_plate.append(row[1:len(row)])
        # if we've seen 4 empty rows, the plate is finished
        # store the plate in the dict under its ID and initialize the next plate
        if empty_row_counter == 4:
            plates[ID] = new_plate
            new_plate = openpyxl.Workbook().create_sheet()  # reset plate to return
            empty_row_counter = 0
        if row[0] is None:
            empty_row_counter += 1
        # Store the last plate after iterating over all the rows
        if row_index == max_row:
            plates[ID] = new_plate
    return plates


# input an openpyxl sheet and return a dictionary with ID as keys and a list of empty or water wells
def process_plate(plate):
    row_names = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    row_index = 0
    empty_wells = []

    # Get the number of rows and columns
    max_row = plate.max_row
    max_column = plate.max_column

    # Iterate over rows and columns using indices
    for row_index in range(1, max_row + 1):  # 1-based index
        for col_index in range(1, max_column + 1):
            cell_value = plate.cell(row=row_index, column=col_index).value
            if cell_value is None or cell_value.strip().lower() == 'water':
                empty_wells.append(row_names[row_index - 1] + str(col_index))
    return empty_wells


def get_empty_wells(path):
    # Select the active worksheet and get the plate maps within
    if not path:
        return None

    empty_wells = {}
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    plates = get_plates(sheet)

    for ID, plate in plates.items():
        empty_wells[ID] = process_plate(plate)

    return empty_wells
